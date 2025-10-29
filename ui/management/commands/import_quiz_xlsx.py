# ui/management/commands/import_quiz_xlsx.py
from django.core.management.base import BaseCommand, CommandError
from ui.models import Lesson, Quiz, Question, Choice
from openpyxl import load_workbook

"""
Format (Sheet1):
A: question
B: option1
C: option2
D: option3
E: option4  (ixtiyoriy)
F: correct  (1..4)
G: order    (ixtiyoriy)

Har satr bitta savol.
"""

class Command(BaseCommand):
    help = "Import quiz questions from XLSX into a Lesson(kind=quiz)."

    def add_arguments(self, parser):
        parser.add_argument("lesson_id", type=int, help="Lesson ID (kind=quiz)")
        parser.add_argument("xlsx_path", type=str, help="Path to .xlsx file")

    def handle(self, *args, **opts):
        lesson_id = opts["lesson_id"]
        path = opts["xlsx_path"]

        lesson = Lesson.objects.filter(pk=lesson_id, kind="quiz").first()
        if not lesson:
            raise CommandError("Lesson not found or not a quiz kind.")

        wb = load_workbook(path)
        ws = wb.active

        quiz, _ = Quiz.objects.get_or_create(lesson=lesson)
        # Eski savollarni o'chirishni xohlasangiz, quyidagini yoqing:
        quiz.questions.all().delete()

        count = 0
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            q_text, o1, o2, o3, o4, correct, q_order = row[0], row[1], row[2], row[3], row[4], row[5], row[6] if len(row) > 6 else None
            if not q_text:
                continue
            q = Question.objects.create(
                quiz=quiz,
                text=str(q_text),
                order=int(q_order) if q_order else (count + 1),
            )
            options = [o1, o2, o3, o4]
            for i, opt in enumerate(options, start=1):
                if opt:
                    Choice.objects.create(
                        question=q,
                        text=str(opt),
                        is_correct=(int(correct) == i if correct else False),
                    )
            count += 1

        self.stdout.write(self.style.SUCCESS(f"Imported {count} questions into lesson {lesson.id}."))
